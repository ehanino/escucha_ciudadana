"""
Management command: cargar_centros_callao
Importa los centros de votación del Callao desde el Excel oficial (EG2026),
actualizando campos existentes en lugar de duplicarlos y cargando el nuevo atributo 'electores'.
"""
import os
import unicodedata
from django.core.management.base import BaseCommand
from django.conf import settings
from personeros.models import CentroVotacion, Distrito


def _limpiar(texto):
    """Elimina espacios extra y normaliza el string."""
    if not texto:
        return ""
    return " ".join(str(texto).strip().split())


def normalizar_distrito(nombre_distrito):
    """Normaliza nombres de distritos para mapear a la BD."""
    if not nombre_distrito:
        return ""
    nombre = nombre_distrito.strip().upper()
    nombre = nombre.replace(" - ", " ")
    nombre = nombre.replace("-", " ")
    return " ".join(nombre.split())


def norm_comparar(s):
    """Normaliza strings para comparaciones difusas (remueve acentos, Ñs, etc)."""
    if not s:
        return ""
    s = "".join(c for c in unicodedata.normalize("NFD", str(s)) if unicodedata.category(c) != "Mn")
    return " ".join(s.upper().replace("Ñ", "N").split())


class Command(BaseCommand):
    help = 'Importa y actualiza centros de votacion del Callao desde el Excel oficial'

    def add_arguments(self, parser):
        parser.add_argument(
            '--flush',
            action='store_true',
            help='Elimina todos los centros de Callao existentes antes de importar (cuidado con referencias)',
        )
        parser.add_argument(
            '--file',
            type=str,
            default=None,
            help='Ruta al archivo Excel (por defecto: docs/Ficha Tcnica EG2026v08 Callao.xlsx)',
        )

    def handle(self, *args, **options):
        import openpyxl

        # Ruta del Excel
        file_path = options['file'] or os.path.join(
            settings.BASE_DIR, 'docs', 'Ficha Tcnica EG2026v08 Callao.xlsx'
        )

        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f'[ERROR] Archivo no encontrado: {file_path}'))
            return

        # Flush opcional (maneja dependencias protegidas)
        if options['flush']:
            callao_centros = CentroVotacion.objects.filter(distrito__provincia__nombre__iexact='CALLAO')
            total_centros = callao_centros.count()
            eliminados = 0
            for c in callao_centros:
                try:
                    c.delete()
                    eliminados += 1
                except Exception as e:
                    self.stderr.write(self.style.WARNING(f'[WARN] No se pudo eliminar centro {c.nombre} (ID: {c.id}) por referencias: {e}'))
            self.stdout.write(self.style.WARNING(f'[INFO] {eliminados} de {total_centros} centros eliminados.'))

        # Cargar libro de Excel
        self.stdout.write(self.style.HTTP_INFO(f'[INFO] Cargando libro de Excel: {file_path}'))
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb['LOCALES'] if 'LOCALES' in wb.sheetnames else wb.active
        self.stdout.write(self.style.HTTP_INFO(f'[INFO] Usando hoja de calculo: {sheet.title}'))

        # Escanear columnas del encabezado dinámicamente
        header_row = None
        col_indices = {}
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if any(isinstance(val, str) and 'NOMBRE DEL LOCAL' in val.upper() for val in row if val):
                header_row = row_idx
                for col_idx, val in enumerate(row):
                    if not val:
                        continue
                    val_upper = str(val).upper().strip()
                    if 'PROVINCIA' in val_upper:
                        col_indices['provincia'] = col_idx
                    elif 'DISTRITO' in val_upper:
                        if '/' in val_upper or 'CIUDAD' in val_upper:
                            col_indices['distrito'] = col_idx
                    elif 'NOMBRE DEL LOCAL' in val_upper:
                        col_indices['nombre'] = col_idx
                    elif 'DIRECCI' in val_upper and 'LOCAL' in val_upper:
                        col_indices['direccion'] = col_idx
                    elif val_upper == 'MESAS':
                        col_indices['actas'] = col_idx
                    elif val_upper == 'ELECTORES':
                        col_indices['electores'] = col_idx
                break

        required_cols = ['provincia', 'distrito', 'nombre', 'direccion', 'actas', 'electores']
        if not header_row or not all(c in col_indices for c in required_cols):
            missing = [c for c in required_cols if c not in col_indices]
            self.stderr.write(self.style.ERROR(
                f'[ERROR] No se pudo encontrar el encabezado o faltan columnas requeridas en el Excel. Faltan: {missing}'
            ))
            return

        self.stdout.write(self.style.SUCCESS(
            f'[INFO] Encabezado detectado en Fila {header_row}. Mapeo de columnas: {col_indices}'
        ))

        creados = 0
        actualizados = 0
        errores = 0
        updated_ids = set()
        used_db_ids = set()

        # Recorrer filas de datos
        for lineno, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not row or len(row) <= max(col_indices.values()):
                continue
            if not any(row):
                continue

            try:
                provincia_raw = row[col_indices['provincia']]
                if not provincia_raw or _limpiar(provincia_raw).upper() != 'CALLAO':
                    continue

                distrito_raw = row[col_indices['distrito']]
                nombre_raw = row[col_indices['nombre']]
                direccion_raw = row[col_indices['direccion']]
                mesas_raw = row[col_indices['actas']]
                electores_raw = row[col_indices['electores']]

                distrito_nom = normalizar_distrito(distrito_raw)
                nombre = _limpiar(nombre_raw).upper()
                direccion = _limpiar(direccion_raw).upper()
                mesas = int(mesas_raw) if mesas_raw is not None else 0
                electores = int(electores_raw) if electores_raw is not None else 0

                if not nombre or not distrito_nom:
                    continue

                # Buscar Distrito en la base de datos
                distrito_obj = Distrito.objects.filter(
                    nombre__iexact=distrito_nom,
                    provincia__nombre__iexact='CALLAO'
                ).first()

                if not distrito_obj:
                    # Intento alternativo sin filtrar provincia en caso de desajuste
                    distrito_obj = Distrito.objects.filter(nombre__iexact=distrito_nom).first()

                if not distrito_obj:
                    self.stderr.write(self.style.WARNING(
                        f'[WARN] Linea {lineno}: Distrito "{distrito_nom}" no encontrado en BD.'
                    ))
                    errores += 1
                    continue

                # Intentar emparejar con un local existente en este distrito (evitando reutilizar IDs)
                obj = None

                # 1. Búsqueda exacta por (nombre, distrito)
                obj = CentroVotacion.objects.filter(
                    distrito=distrito_obj,
                    nombre__iexact=nombre
                ).exclude(id__in=used_db_ids).first()

                # 2. Búsqueda por similitud de nombre
                if not obj:
                    norm_nombre = norm_comparar(nombre)
                    for c in CentroVotacion.objects.filter(distrito=distrito_obj).exclude(id__in=used_db_ids):
                        c_norm = norm_comparar(c.nombre)
                        if c_norm == norm_nombre or norm_nombre in c_norm or c_norm in norm_nombre:
                            obj = c
                            break

                # 3. Búsqueda por similitud de dirección
                if not obj and direccion:
                    norm_dir = norm_comparar(direccion)
                    for c in CentroVotacion.objects.filter(distrito=distrito_obj).exclude(id__in=used_db_ids):
                        if c.direccion:
                            c_dir_norm = norm_comparar(c.direccion)
                            if norm_dir in c_dir_norm or c_dir_norm in norm_dir:
                                obj = c
                                break

                # Actualizar o Crear
                if obj:
                    # Actualizar en lugar
                    obj.nombre = nombre
                    obj.direccion = direccion
                    obj.actas = mesas
                    obj.electores = electores
                    obj.save()
                    actualizados += 1
                else:
                    # Crear nuevo (respetando restricciones únicas)
                    obj, created = CentroVotacion.objects.update_or_create(
                        nombre=nombre,
                        distrito=distrito_obj,
                        defaults={
                            'direccion': direccion,
                            'actas': mesas,
                            'electores': electores
                        }
                    )
                    if created:
                        creados += 1
                    else:
                        actualizados += 1

                updated_ids.add(obj.id)
                used_db_ids.add(obj.id)

            except Exception as exc:
                errores += 1
                self.stderr.write(self.style.WARNING(f'[WARN] Linea {lineno} ignorada: {exc}'))

        # Eliminar locales obsoletos en Callao que no se encontraron en la hoja de cálculo
        obsolete_qs = CentroVotacion.objects.filter(
            distrito__provincia__nombre__iexact='CALLAO'
        ).exclude(id__in=updated_ids)

        eliminados = 0
        for obs in obsolete_qs:
            try:
                obs.delete()
                eliminados += 1
            except Exception as e:
                self.stderr.write(self.style.WARNING(
                    f'[WARN] No se pudo eliminar local obsoleto "{obs.nombre}" (ID: {obs.id}) por referencias: {e}'
                ))

        # Resumen final
        self.stdout.write(self.style.SUCCESS(
            f'\nImportacion finalizada:\n'
            f'   Centros creados    : {creados}\n'
            f'   Centros actualiz.  : {actualizados}\n'
            f'   Centros eliminados : {eliminados}\n'
            f'   Errores            : {errores}\n'
            f'   Total en BD        : {CentroVotacion.objects.count()}'
        ))
