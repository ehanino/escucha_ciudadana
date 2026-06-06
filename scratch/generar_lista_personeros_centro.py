import os
import sys
import django

# Añadir la raíz del proyecto al path
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(PROJECT_ROOT)

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'escucha_ciudadana.settings')
django.setup()

from personeros.models import Personero
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_report():
    print("Iniciando la extracción de personeros de centro de votación...")
    
    # 1. Consultar todos los personeros con cargo Coordinador2 (Personero de Centro de Votación)
    # sin filtro de estado, ordenando por distrito y nombres
    personeros = Personero.objects.filter(cargo='Coordinador2').select_related('distrito').order_by(
        'distrito__nombre', 'apellido_paterno', 'apellido_materno'
    )
    
    # 2. Inicializar el workbook de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Personeros de Centro"
    
    # Mostrar líneas de cuadrícula
    ws.views.sheetView[0].showGridLines = True
    
    # Cabeceras
    headers = [
        "Distrito",
        "Apellido Paterno",
        "Apellido Materno",
        "Nombres",
        "DNI",
        "Nro. Celular"
    ]
    ws.append(headers)
    
    # Estilos
    header_fill = PatternFill(start_color="39B54A", end_color="39B54A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Aplicar estilos a la cabecera
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    # 3. Rellenar datos de personeros
    row_idx = 2
    for p in personeros:
        distrito_nombre = p.distrito.nombre if p.distrito else "SIN DISTRITO"
        
        ws.cell(row=row_idx, column=1, value=distrito_nombre).alignment = left_align
        ws.cell(row=row_idx, column=2, value=p.apellido_paterno).alignment = left_align
        ws.cell(row=row_idx, column=3, value=p.apellido_materno).alignment = left_align
        ws.cell(row=row_idx, column=4, value=p.nombres).alignment = left_align
        ws.cell(row=row_idx, column=5, value=p.dni).alignment = center_align
        ws.cell(row=row_idx, column=6, value=p.nro_celular).alignment = center_align
        
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            
        row_idx += 1
        
    # Auto-ajustar el ancho de las columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Guardar en la raíz
    output_path = os.path.join(PROJECT_ROOT, "reporte_personeros_centro.xlsx")
    wb.save(output_path)
    print(f"Reporte generado exitosamente en: {output_path}")

if __name__ == "__main__":
    generate_report()
