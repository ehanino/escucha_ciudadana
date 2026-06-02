import os
import sys
import django

# Añadir la raíz del proyecto al path
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(PROJECT_ROOT)

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'escucha_ciudadana.settings')
django.setup()

from personeros.models import CentroVotacion, Personero
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_report():
    print("Iniciando la generación del reporte Excel...")
    
    # 1. Obtener los centros de votación en Callao (Dep: 07, Prov: 0701)
    # Filtrar solo aquellos que tengan al menos un personero 'confirmado' de cualquier tipo
    centros = CentroVotacion.objects.filter(
        distrito__provincia__departamento_id='07',
        distrito__provincia_id='0701'
    ).select_related('distrito')
    
    # Filtrar centros que contengan personeros confirmados
    centros_con_personeros = centros.filter(
        personeros__estado='confirmado'
    ).distinct()
    
    # Ordenar por Distrito (alfabéticamente) y luego por nombre de Centro de Votación
    centros_con_personeros = centros_con_personeros.order_by('distrito__nombre', 'nombre')
    
    # 2. Crear el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Locales Callao"
    
    # Mostrar las líneas de cuadrícula
    ws.views.sheetView[0].showGridLines = True
    
    # Cabeceras
    headers = [
        "Distrito",
        "Centro de Votación",
        "Coordinador Local",
        "Personero de Centro de Votación",
        "Número de Mesas",
        "Total de Personeros"
    ]
    ws.append(headers)
    
    # Estilos
    header_fill = PatternFill(start_color="39B54A", end_color="39B54A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Aplicar estilos a la cabecera
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # 3. Llenar los datos de cada local
    row_idx = 2
    for cv in centros_con_personeros:
        # Obtener el Personero de Centro de Votación (cargo 'Coordinador2' y estado 'confirmado')
        p_centro = cv.personeros.filter(cargo='Coordinador2', estado='confirmado').first()
        p_centro_nombre = p_centro.nombre_completo if p_centro else ""
        
        # Conteo de Personeros de Mesa: únicamente cargo='Coordinador3' y estado='confirmado'
        total_personeros_mesa = cv.personeros.filter(cargo='Coordinador3', estado='confirmado').count()
        
        distrito_nombre = cv.distrito.nombre if cv.distrito else "SIN DISTRITO"
        
        # Escribir valores
        ws.cell(row=row_idx, column=1, value=distrito_nombre).alignment = left_align
        ws.cell(row=row_idx, column=2, value=cv.nombre).alignment = left_align
        ws.cell(row=row_idx, column=3, value="").alignment = left_align # Coordinador Local se deja vacío
        ws.cell(row=row_idx, column=4, value=p_centro_nombre).alignment = left_align
        ws.cell(row=row_idx, column=5, value=cv.actas).alignment = right_align
        ws.cell(row=row_idx, column=6, value=total_personeros_mesa).alignment = right_align
        
        # Aplicar formato de bordes y fuente a las celdas de datos
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            
        row_idx += 1
        
    # Auto-ajustar el ancho de las columnas según su contenido
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Definir ruta de salida en la raíz del proyecto
    output_path = os.path.join(PROJECT_ROOT, "reporte_personeros_callao.xlsx")
    wb.save(output_path)
    print(f"Reporte generado exitosamente en: {output_path}")

if __name__ == "__main__":
    generate_report()
